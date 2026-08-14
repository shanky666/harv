import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

OUT = r"C:\Users\madha\AppData\Local\Temp\opencode\harv\HarvestLenz_dataset_sources.xlsx"

HEADERS = ["Item", "Source link"]

ROWS = [
    # Fruits (15; mango/pineapple/grapes/pomegranate excluded)
    ["Orange", "https://data.mendeley.com/datasets/3f83gxmv57/2"],
    ["Guava", "https://data.mendeley.com/datasets/w3fg8jjmzr/1"],
    ["Kiwi", "https://www.kaggle.com/datasets/moltean/fruits"],
    ["Watermelon", "https://www.kaggle.com/datasets/parvezthabarak/freshness50-food-classification"],
    ["Banana", "https://www.kaggle.com/datasets/sriramr/fruits-fresh-and-rotten-for-classification"],
    ["Cocoa", "https://data.mendeley.com/datasets/sr279sf4hs/1"],
    ["Coffee", "https://www.kaggle.com/datasets/sujitraarw/coffee-green-bean-with-17-defects-original"],
    ["Strawberry", "https://www.kaggle.com/datasets/muhammad0subhan/fruit-and-vegetable-disease-healthy-vs-rotten"],
    ["Plum", "https://huggingface.co/datasets/Project-AgML/african_plum_grading_classification"],
    ["Peach", "https://zenodo.org/records/7224690"],
    ["Pear", "https://zenodo.org/records/7224690"],
    # Vegetables (5)
    ["Carrot", "https://www.kaggle.com/datasets/muhammad0subhan/fruit-and-vegetable-disease-healthy-vs-rotten"],
    ["Tomato", "https://data.mendeley.com/datasets/42m5tb7yv9/1"],
    ["Cucumber", "https://data.mendeley.com/datasets/42m5tb7yv9/1"],
    ["Capsicum", "https://www.kaggle.com/datasets/muhriddinmuxiddinov/fruits-and-vegetables-dataset"],
    ["Potato", "https://data.mendeley.com/datasets/7vm7xskfg4/1"],
]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sources"

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")

for c, h in enumerate(HEADERS, start=1):
    cell = ws.cell(row=1, column=c, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill

for r, (item, link) in enumerate(ROWS, start=2):
    ws.cell(row=r, column=1, value=item)
    lc = ws.cell(row=r, column=2, value=link)
    lc.hyperlink = link
    lc.style = "Hyperlink"

ws.column_dimensions["A"].width = 16
ws.column_dimensions["B"].width = 80
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:B{len(ROWS)+1}"

wb.save(OUT)
print("WROTE", OUT, "| items:", len(ROWS))
